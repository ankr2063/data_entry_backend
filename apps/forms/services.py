import requests
from msal import ConfidentialClientApplication
from typing import Dict, List, Any
from decouple import config
from django.db import transaction
from .models import Form, FormDisplayVersion, FormEntryVersion
# import concurrent.futures  # No longer needed - was used for Graph API batch processing
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from io import BytesIO
import zipfile
import xml.etree.ElementTree as ET


class SharePointService:
    def __init__(self):
        self.client_app = ConfidentialClientApplication(
            config('MICROSOFT_CLIENT_ID'),
            authority=f"https://login.microsoftonline.com/{config('MICROSOFT_TENANT_ID')}",
            client_credential=config('MICROSOFT_CLIENT_SECRET'),
        )
        self.token = None
    
    def get_access_token(self) -> str:
        if not self.token:
            result = self.client_app.acquire_token_for_client(
                scopes=["https://graph.microsoft.com/.default"]
            )
            if "access_token" in result:
                self.token = result["access_token"]
            else:
                raise Exception(f"Failed to acquire token: {result.get('error_description')}")
        return self.token
    
    def create_new_form(self, sharepoint_url: str, form_name: str, created_by: str, updated_by: str) -> Dict:
        """Create new form from SharePoint URL"""
        worksheets = self.get_workbook_worksheets(sharepoint_url)
        
        display_sheet = None
        entry_sheet = None
        
        for worksheet in worksheets:
            name = worksheet['name'].lower()
            if 'display' in name:
                display_sheet = worksheet
            elif 'entry' in name:
                entry_sheet = worksheet
        
        if not display_sheet or not entry_sheet:
            raise Exception("Both 'display' and 'entry' worksheets are required")
        
        with transaction.atomic():
            form = Form.objects.create(
                form_name=form_name,
                source='sharepoint',
                url=sharepoint_url,
                created_by=created_by,
                updated_by=updated_by
            )
            
            display_metadata = self.get_display_sheet_metadata(sharepoint_url, display_sheet['name'])
            
            FormDisplayVersion.objects.create(
                form=form,
                form_display_json=display_metadata,
                form_version='1',
                approved=False,
                created_by=created_by,
                updated_by=updated_by
            )
            
            entry_data = self.get_entry_sheet_data(sharepoint_url, entry_sheet['name'])
            
            FormEntryVersion.objects.create(
                form=form,
                form_entry_json=entry_data,
                form_version='1',
                approved=False,
                created_by=created_by,
                updated_by=updated_by
            )
            
            return {
                'form_id': form.id,
                'form_name': form_name,
                'display_version': 1,
                'entry_version': 1,
                'display_sheet': display_sheet['name'],
                'entry_sheet': entry_sheet['name']
            }
    
    def update_existing_form(self, form_id: int, updated_by: str) -> Dict:
        """Update existing form from SharePoint URL"""
        form = Form.objects.get(id=form_id)
        sharepoint_url = form.url
        worksheets = self.get_workbook_worksheets(sharepoint_url)
        
        display_sheet = None
        entry_sheet = None
        
        for worksheet in worksheets:
            name = worksheet['name'].lower()
            if 'display' in name:
                display_sheet = worksheet
            elif 'entry' in name:
                entry_sheet = worksheet
        
        if not display_sheet or not entry_sheet:
            raise Exception("Both 'display' and 'entry' worksheets are required")
        
        new_display_metadata = self.get_display_sheet_metadata(sharepoint_url, display_sheet['name'])
        new_entry_data = self.get_entry_sheet_data(sharepoint_url, entry_sheet['name'])
        
        latest_display = FormDisplayVersion.objects.filter(form=form).order_by('-form_version').first()
        latest_entry = FormEntryVersion.objects.filter(form=form).order_by('-form_version').first()
        
        versions_updated = []
        display_version = int(latest_display.form_version) if latest_display else 0
        entry_version = int(latest_entry.form_version) if latest_entry else 0
        
        with transaction.atomic():
            form.url = sharepoint_url
            form.updated_by = updated_by
            form.save()
            
            if not latest_display or latest_display.form_display_json != new_display_metadata:
                display_version += 1
                FormDisplayVersion.objects.create(
                    form=form,
                    form_display_json=new_display_metadata,
                    form_version=str(display_version),
                    approved=False,
                    created_by=updated_by,
                    updated_by=updated_by
                )
                versions_updated.append('display')
            
            if not latest_entry or latest_entry.form_entry_json != new_entry_data:
                entry_version += 1
                FormEntryVersion.objects.create(
                    form=form,
                    form_entry_json=new_entry_data,
                    form_version=str(entry_version),
                    approved=False,
                    created_by=updated_by,
                    updated_by=updated_by
                )
                versions_updated.append('entry')
            
            return {
                'form_id': form.id,
                'form_name': form.form_name,
                'display_version': display_version,
                'entry_version': entry_version,
                'versions_updated': versions_updated,
                'display_sheet': display_sheet['name'],
                'entry_sheet': entry_sheet['name']
            }
    
    def get_display_sheet_metadata(self, sharepoint_url: str, worksheet_name: str) -> Dict:
        """Get complete metadata for display sheet using openpyxl"""
        return self._get_display_metadata_from_file(sharepoint_url, worksheet_name)
    
    def get_entry_sheet_data(self, sharepoint_url: str, worksheet_name: str) -> List[Dict]:
        """Get entry sheet data and transform to JSON object list"""
        token = self.get_access_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        site_id, file_path = self._parse_sharepoint_url(sharepoint_url)
        base_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{file_path}:/workbook/worksheets('{worksheet_name}')"
        
        range_response = requests.get(f"{base_url}/usedRange", headers=headers)
        if range_response.status_code != 200:
            raise Exception(f"Failed to get range: {range_response.text}")
        
        range_data = range_response.json()
        values = range_data.get("values", [])
        
        if not values or len(values) < 2:
            return []
        
        headers_row = values[0]
        data_rows = values[1:]
        
        json_objects = []
        for row in data_rows:
            row_obj = {}
            for i, header in enumerate(headers_row):
                if i < len(row):
                    row_obj[str(header)] = row[i]
                else:
                    row_obj[str(header)] = None
            json_objects.append(row_obj)
        
        return json_objects
    
    def get_workbook_worksheets(self, sharepoint_url: str) -> List[Dict]:
        token = self.get_access_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        site_id, file_path = self._parse_sharepoint_url(sharepoint_url)
        url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{file_path}:/workbook/worksheets"
        
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            return response.json().get("value", [])
        else:
            raise Exception(f"Failed to get worksheets: {response.text}")
    
    def _parse_sharepoint_url(self, sharepoint_url: str) -> tuple:
        """Parse SharePoint sharing URL to extract site URL and file path"""
        import re
        from urllib.parse import urlparse, parse_qs, unquote
        
        parsed = urlparse(sharepoint_url)
        domain = parsed.netloc
        path_parts = parsed.path.split('/')
        
        # Check if it's a Doc.aspx URL format
        if 'Doc.aspx' in sharepoint_url or '_layouts' in sharepoint_url:
            # Extract site name from /sites/xxx/
            site_name = None
            for i, part in enumerate(path_parts):
                if part == 'sites' and i + 1 < len(path_parts):
                    site_name = path_parts[i + 1]
                    break
            
            if not site_name:
                raise Exception("Could not extract site name from URL")
            
            # Extract file GUID from sourcedoc parameter
            query_params = parse_qs(parsed.query)
            source_doc = query_params.get('sourcedoc', [None])[0]
            file_name = query_params.get('file', [None])[0]
            
            if not source_doc or not file_name:
                raise Exception("Could not extract file info from URL")
            
            # Remove curly braces from GUID
            file_guid = source_doc.strip('{}').replace('%7B', '').replace('%7D', '')
            
            site_url = f"https://{domain}/sites/{site_name}"
            site_id = self._get_site_id(site_url)
            
            return site_id, file_name
        
        # Original sharing link format (:x:/s/test/...)
        else:
            # Extract site name (after /s/)
            site_name = None
            for i, part in enumerate(path_parts):
                if part == 's' and i + 1 < len(path_parts):
                    site_name = path_parts[i + 1]
                    break
            
            if not site_name:
                raise Exception("Could not extract site name from URL")
            
            site_url = f"https://{domain}/sites/{site_name}"
            site_id = self._get_site_id(site_url)
            
            # Extract file ID from URL path
            file_id_match = re.search(r'/([A-Za-z0-9_-]+)\?', sharepoint_url)
            if not file_id_match:
                raise Exception("Could not extract file ID from URL")
            
            file_id = file_id_match.group(1)
            file_path = self._get_file_path_from_id(site_id, file_id)
            
            return site_id, file_path
    
    def _get_file_path_from_id(self, site_id: str, file_id: str) -> str:
        """Get file path from sharing link ID"""
        token = self.get_access_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        # Try to resolve the sharing link
        sharing_url = f"https://graph.microsoft.com/v1.0/shares/u!{file_id}/driveItem"
        response = requests.get(sharing_url, headers=headers)
        
        if response.status_code == 200:
            item = response.json()
            return item.get('name', 'Unknown.xlsx')
        else:
            # Fallback: search for files in the site
            search_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children"
            response = requests.get(search_url, headers=headers)
            if response.status_code == 200:
                items = response.json().get('value', [])
                # Return first Excel file found
                for item in items:
                    if item.get('name', '').endswith(('.xlsx', '.xls')):
                        return item['name']
            
            raise Exception("Could not resolve file path from sharing URL")
    
    def _get_site_id(self, site_url: str) -> str:
        token = self.get_access_token()
        headers = {"Authorization": f"Bearer {token}"}
        
        parts = site_url.replace("https://", "").split("/")
        domain = parts[0]
        site_path = "/".join(parts[1:])
        
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{domain}:/{site_path}"
        response = requests.get(graph_url, headers=headers)
        
        if response.status_code == 200:
            return response.json()["id"]
        else:
            raise Exception(f"Failed to get site ID: {response.text}")

    
    def _get_display_metadata_from_file(self, sharepoint_url: str, worksheet_name: str) -> Dict:
        """Extract complete display metadata using openpyxl"""
        try:
            token = self.get_access_token()
            headers = {"Authorization": f"Bearer {token}"}
            
            site_id, file_path = self._parse_sharepoint_url(sharepoint_url)
            download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{file_path}:/content"
            
            response = requests.get(download_url, headers=headers)
            if response.status_code != 200:
                raise Exception(f"Failed to download file: {response.status_code}")
            
            file_content = BytesIO(response.content)
            
            # Extract theme colors from Excel file
            theme_colors = self._extract_theme_colors(file_content)
            file_content.seek(0)  # Reset file pointer
            
            wb = load_workbook(file_content, data_only=False)
            ws = wb[worksheet_name]
            
            # Load workbook again with data_only=True to get calculated values
            file_content.seek(0)
            wb_data = load_workbook(file_content, data_only=True)
            ws_data = wb_data[worksheet_name]
            
            # Determine dimensions - find actual used columns
            max_row = ws.max_row
            
            # Find actual used columns (non-empty cells)
            used_cols = set()
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        used_cols.add(col_idx)
            
            actual_max_col = max(used_cols) if used_cols else ws.max_column
            
            # Extract cells metadata
            cells_metadata = []
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, actual_max_col + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell_data_value = ws_data.cell(row_idx, col_idx)
                    cell_data = self._extract_openpyxl_cell_data(cell, cell_data_value, row_idx - 1, col_idx - 1, ws, theme_colors)
                    cells_metadata.append(cell_data)
            
            # Extract merged cells
            merged_cells = []
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_col <= actual_max_col:
                    merged_cells.append({
                        "range": str(merged_range),
                        "start_row": merged_range.min_row - 1,
                        "start_col": merged_range.min_col - 1,
                        "row_span": merged_range.max_row - merged_range.min_row + 1,
                        "col_span": merged_range.max_col - merged_range.min_col + 1
                    })
            
            return {
                "worksheet_name": worksheet_name,
                "dimensions": {"rows": max_row, "columns": actual_max_col},
                "cells": cells_metadata,
                "merged_cells": merged_cells
            }
            
        except Exception as e:
            raise Exception(f"Failed to extract display metadata: {e}")
    
    def _extract_openpyxl_cell_data(self, cell, cell_data_value, row: int, col: int, ws, theme_colors: Dict) -> Dict:
        """Extract all metadata from an openpyxl cell"""
        # Handle merged cells - they don't have full attributes
        if isinstance(cell, MergedCell):
            col_letter = chr(65 + col) if col < 26 else chr(64 + col // 26) + chr(65 + col % 26)
            return {
                "address": f"{col_letter}{row + 1}",
                "value": None,
                "formula": None,
                "display_value": "",
                "data_type": "empty",
                "row": row,
                "column": col,
                "font": {"name": "", "size": 11, "bold": False, "italic": False, "underline": "none", "strikethrough": False, "color": ""},
                "fill": {"color": "", "pattern_type": ""},
                "alignment": {"horizontal": "", "vertical": "", "wrap_text": False, "indent": 0, "text_rotation": 0},
                "borders": {"left": {"style": ""}, "right": {"style": ""}, "top": {"style": ""}, "bottom": {"style": ""}},
                "number_format": {"format": ""},
                "protection": {"locked": True},
                "column_width": ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions else 8.43,
                "row_height": ws.row_dimensions[row + 1].height if row + 1 in ws.row_dimensions else 15,
                "column_hidden": False,
                "row_hidden": False,
                "hyperlink": {"target": ""},
                "comment": {"text": ""},
                "validation": {},
                "conditional_formats": []
            }
        
        # Separate formula and value
        formula_value = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
        display_value = cell_data_value.value if cell_data_value else cell.value
        
        # Convert datetime to string
        if hasattr(display_value, 'isoformat'):
            display_value = display_value.isoformat()
        elif display_value is not None and not isinstance(display_value, (str, int, float, bool)):
            display_value = str(display_value)
        
        return {
            "address": cell.coordinate,
            "value": display_value,
            "formula": formula_value,
            "display_value": str(display_value) if display_value is not None else "",
            "data_type": self._infer_data_type(display_value),
            "row": row,
            "column": col,
            "font": {
                "name": cell.font.name if cell.font else "",
                "size": cell.font.size if cell.font else 11,
                "bold": cell.font.bold if cell.font else False,
                "italic": cell.font.italic if cell.font else False,
                "underline": cell.font.underline if cell.font else "none",
                "strikethrough": cell.font.strike if cell.font else False,
                "color": self._get_color_value(cell.font.color, theme_colors) if cell.font and cell.font.color else ""
            },
            "fill": {
                "color": self._get_color_value(cell.fill.fgColor, theme_colors) if cell.fill and cell.fill.fgColor and cell.fill.patternType and cell.fill.patternType != 'none' else "",
                "pattern_type": cell.fill.patternType if cell.fill else ""
            },
            "alignment": {
                "horizontal": cell.alignment.horizontal if cell.alignment else "",
                "vertical": cell.alignment.vertical if cell.alignment else "",
                "wrap_text": cell.alignment.wrap_text if cell.alignment else False,
                "indent": cell.alignment.indent if cell.alignment else 0,
                "text_rotation": cell.alignment.text_rotation if cell.alignment else 0
            },
            "borders": {
                "left": {"style": cell.border.left.style if cell.border and cell.border.left else ""},
                "right": {"style": cell.border.right.style if cell.border and cell.border.right else ""},
                "top": {"style": cell.border.top.style if cell.border and cell.border.top else ""},
                "bottom": {"style": cell.border.bottom.style if cell.border and cell.border.bottom else ""}
            },
            "number_format": {
                "format": cell.number_format if cell.number_format else ""
            },
            "protection": {
                "locked": cell.protection.locked if cell.protection else True
            },
            "column_width": ws.column_dimensions[cell.column_letter].width if cell.column_letter in ws.column_dimensions else 8.43,
            "row_height": ws.row_dimensions[cell.row].height if cell.row in ws.row_dimensions else 15,
            "column_hidden": ws.column_dimensions[cell.column_letter].hidden if cell.column_letter in ws.column_dimensions else False,
            "row_hidden": ws.row_dimensions[cell.row].hidden if cell.row in ws.row_dimensions else False,
            "hyperlink": {"target": cell.hyperlink.target if cell.hyperlink else ""},
            "comment": {"text": cell.comment.text if cell.comment else ""},
            "validation": {},
            "conditional_formats": []
        }
    

    
    def _extract_theme_colors(self, file_content: BytesIO) -> Dict:
        """Extract theme colors from Excel file"""
        theme_colors = {}
        try:
            with zipfile.ZipFile(file_content, 'r') as zip_ref:
                theme_xml = zip_ref.read('xl/theme/theme1.xml')
                root = ET.fromstring(theme_xml)
                
                # Define namespace
                ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
                
                # Extract color scheme
                clr_scheme = root.find('.//a:clrScheme', ns)
                if clr_scheme:
                    color_map = {
                        'lt1': 0, 'dk1': 1, 'lt2': 2, 'dk2': 3,
                        'accent1': 4, 'accent2': 5, 'accent3': 6,
                        'accent4': 7, 'accent5': 8, 'accent6': 9
                    }
                    
                    for color_name, theme_idx in color_map.items():
                        color_elem = clr_scheme.find(f'.//a:{color_name}', ns)
                        if color_elem:
                            # Check for srgbClr or sysClr
                            srgb = color_elem.find('.//a:srgbClr', ns)
                            if srgb is not None and 'val' in srgb.attrib:
                                theme_colors[theme_idx] = srgb.attrib['val']
                            else:
                                sys_clr = color_elem.find('.//a:sysClr', ns)
                                if sys_clr is not None and 'lastClr' in sys_clr.attrib:
                                    theme_colors[theme_idx] = sys_clr.attrib['lastClr']
        except:
            pass
        
        return theme_colors

    def _get_color_value(self, color_obj, theme_colors: Dict) -> str:
        """Extract color value from openpyxl color object"""
        if not color_obj:
            return ""
        
        try:
            # Check RGB first
            if hasattr(color_obj, 'rgb') and color_obj.rgb and isinstance(color_obj.rgb, str):
                rgb = color_obj.rgb
                if rgb not in ['00000000', 'FF000000']:
                    return rgb[2:] if len(rgb) == 8 else rgb
            
            # Handle theme colors
            if hasattr(color_obj, 'theme') and color_obj.theme is not None:
                base_color = theme_colors.get(color_obj.theme, '')
                if base_color:
                    # Apply tint if present
                    if hasattr(color_obj, 'tint') and color_obj.tint != 0:
                        return self._apply_tint(base_color, color_obj.tint)
                    return base_color
        except:
            pass
        
        return ""
    
    def _apply_tint(self, rgb: str, tint: float) -> str:
        """Apply tint to RGB color"""
        try:
            r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
            
            if tint < 0:
                # Darken
                r = int(r * (1 + tint))
                g = int(g * (1 + tint))
                b = int(b * (1 + tint))
            else:
                # Lighten
                r = int(r + (255 - r) * tint)
                g = int(g + (255 - g) * tint)
                b = int(b + (255 - b) * tint)
            
            return f"{r:02X}{g:02X}{b:02X}"
        except:
            return rgb
    
    def _infer_data_type(self, value) -> str:
        if value is None or value == "":
            return "empty"
        elif isinstance(value, bool):
            return "boolean"
        elif isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, str):
            if value.startswith("="):
                return "formula"
            return "text"
        else:
            return "unknown"